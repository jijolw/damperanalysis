from flask import Blueprint, request, jsonify, render_template, send_from_directory, session
import pandas as pd
import os
from .utils import get_google_sheets_data
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

age_analysis_bp = Blueprint("age_analysis_bp", __name__)
REPORT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reports")

@age_analysis_bp.route("/analysis_age")
def analysis_age_page():
    return render_template("analysis_age.html")

@age_analysis_bp.route("/download/reports/<filename>")
def download_report(filename):
    return send_from_directory(REPORT_FOLDER, filename, as_attachment=True)

@age_analysis_bp.route("/analyze_age")
def analyze_age_ajax():
    age_group = request.args.get("value", "").strip()

    try:
        df, error = get_google_sheets_data()
        if error:
            return jsonify({"error": error}), 500

        if df.empty or "Age" not in df.columns:
            return jsonify({"error": "No valid data found!"}), 404

        # Apply date filtering logic from session
        start_date = session.get("start_date")
        end_date = session.get("end_date")
        
        if start_date and end_date:
            try:
                # Ensure dates are in the correct format
                start_date = datetime.strptime(start_date, "%Y-%m-%d")
                end_date = datetime.strptime(end_date, "%Y-%m-%d")

                # Convert 'Test date time' column to datetime
                df["Test date time"] = pd.to_datetime(df["Test date time"], errors='coerce')

                # Filter based on the date range
                df = df[(df["Test date time"] >= start_date) & (df["Test date time"] <= end_date)]
            except Exception as e:
                return jsonify({"error": f"Invalid date range: {e}"}), 400

        df["Age"] = df["Age"].astype(str).str.replace(" days", "", regex=True).str.strip()
        df["Age"] = pd.to_numeric(df["Age"], errors="coerce")

        def categorize_age(days):
            if pd.isna(days): return "3-5 years"
            days = int(days)
            if days < 730: return "Less than 2 years"
            elif days < 1095: return "2-3 years"
            elif days < 1825: return "3-5 years"
            return "Above 5 years"

        df["Age Group"] = df["Age"].apply(categorize_age)
        df_filtered = df[df["Age Group"].str.lower() == age_group.lower()].copy()

        if df_filtered.empty:
            return jsonify({"error": f"No data found for Age Group: {age_group}"}), 404

        df_filtered.loc[:, "Failed"] = (df_filtered["Test Result"] == "FAIL").astype(int)
        df_filtered.loc[:, "Total"] = 1

        summary = df_filtered.groupby(["TYPE OF DAMPER", "Make"]).agg(
            Failures=('Failed', 'sum'),
            Total_Receipts=('Total', 'sum')
        ).reset_index()

        pivot = summary.pivot(index="TYPE OF DAMPER", columns="Make", values=["Failures", "Total_Receipts"]).fillna(0)
        make_order = ["KONI", "SACHS", "KNORR", "IAI", "ESCORTS", "GABRIEL", "SABOHEMA", "OTHER"]

        final = pd.DataFrame(index=pivot.index)
        final.index.name = "Damper Type"

        for make in make_order:
            fail = pivot["Failures"].get(make, pd.Series(0, index=final.index))
            total = pivot["Total_Receipts"].get(make, pd.Series(0, index=final.index))
            percent = (fail / total * 100).replace([float("inf"), -float("inf")], 0).fillna(0)
            final[f"{make} Fail"] = fail.astype(int)
            final[f"{make} Total"] = total.astype(int)
            final[f"{make} %"] = percent.map("{:.2f}".format)

        total_fail = summary.groupby("TYPE OF DAMPER")["Failures"].sum()
        total_recv = summary.groupby("TYPE OF DAMPER")["Total_Receipts"].sum()
        total_percent = (total_fail / total_recv * 100).replace([float("inf"), -float("inf")], 0).fillna(0)

        final["Total Fail"] = total_fail
        final["Total Received"] = total_recv
        final["Total %"] = total_percent.map("{:.2f}".format)

        total_row = pd.DataFrame(index=["Total (All Types)"])
        for make in make_order:
            fail = summary[summary["Make"] == make]["Failures"].sum()
            total = summary[summary["Make"] == make]["Total_Receipts"].sum()
            percent = (fail / total * 100) if total > 0 else 0
            total_row[f"{make} Fail"] = fail
            total_row[f"{make} Total"] = total
            total_row[f"{make} %"] = f"{percent:.2f}"

        total_row["Total Fail"] = total_fail.sum()
        total_row["Total Received"] = total_recv.sum()
        total_row["Total %"] = f"{(total_fail.sum() / total_recv.sum() * 100):.2f}" if total_recv.sum() > 0 else "0.00"
        final = pd.concat([final, total_row])

        # Ensure reports directory exists
        os.makedirs(REPORT_FOLDER, exist_ok=True)

        safe_group = age_group.replace(" ", "_").replace("<", "less_than").replace(">", "greater_than")
        report_filename = f"Age_Analysis_{safe_group}.xlsx"
        report_path = os.path.join(REPORT_FOLDER, report_filename)

        try:
            with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                final.to_excel(writer, sheet_name='Age Analysis', startrow=1)  # Leave first row for heading

                worksheet = writer.sheets['Age Analysis']
                max_col = len(final.columns) + 1
                worksheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=max_col)
                heading_cell = worksheet.cell(row=1, column=2)
                heading_cell.value = f"Failure Summary for Age Group: {age_group}"
                heading_cell.font = Font(bold=True, size=14)

                for i, column in enumerate(final.columns):
                    column_width = max(final[column].astype(str).map(len).max(), len(str(column)))
                    worksheet.column_dimensions[get_column_letter(i + 2)].width = column_width + 2

                for cell in worksheet[2]:
                    cell.font = Font(bold=True)

        except Exception as excel_error:
            print(f"Excel formatting error: {str(excel_error)}. Using basic Excel export.")
            final.to_excel(report_path)

        table_html = f"""
        <h4><strong>Failure Summary for Age Group: {age_group}</strong></h4>
        """ + final.to_html(
            classes="table table-bordered table-striped",
            escape=False,
            na_rep="-"
        )

        download_link = f"/download/reports/{report_filename}"

        return jsonify({
            "table_html": table_html,
            "download_link": download_link
        })

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"error": f"Internal Error: {str(e)}"}), 500
